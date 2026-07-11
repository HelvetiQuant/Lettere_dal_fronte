import threading
import json
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, HTTPException, Body, UploadFile, File, Form
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse

from config import IMI_PDFS, COLUMNS
from database import (
    init_db, get_all_progress, count_internati, get_internati,
    export_excel, export_csv, update_internato, delete_internato,
    get_internato_by_id, count_needs_review, delete_letter, get_latest_active_progress,
    count_fondi, count_menzioni, search_all, get_fondi_summary, delete_fondo_file,
    count_decorati, get_decorati_albi,
    count_entita, count_collegamenti, search_entita, get_collegamenti_entita,
    get_ai_ricerche,
)
import fondi
import decorati
import linker
import ai_research
import caduti_albooro
import caduti_bologna
import caduti_cwgc
import caduti_francia_ww1
import caduti_ministero
import caduti_sardi
import decorati_nastroazzurro
import nara_t315_ocr
import nara_catalog
import archivio_fonti
import memory_router
import source_locator
import soldier_dashboard
import biography
from source_providers.federation import (
    list_providers, get_provider, federated_search,
    fetch_source as fed_fetch_source, get_federation_stats,
)
from downloader import is_downloaded, download_letter, get_downloaded_letters
from extractor import (extract_letter, request_stop, clear_stop_request,
                      is_stop_requested, get_last_completed_letter, clear_last_completed_letter)
from scraper import run_scrape, get_scrape_status
from file_importer import (
    import_csv, import_excel, import_pdf, import_image,
    detect_file_type, _suggest_mapping, UPLOAD_DIR,
)
from geocoder import validate_place, validate_record_locations
from credits import get_usage_summary, init_usage_table

app = FastAPI(title="IMI Extractor - Internati Militari Italiani", version="1.0.0")

_extraction_lock = threading.Lock()
_running_letter = None
_fondi_lock = threading.Lock()
_running_fondo = None
_decorati_lock = threading.Lock()
_linker_lock = threading.Lock()
_albooro_lock = threading.Lock()
_bologna_lock = threading.Lock()
_cwgc_lock = threading.Lock()
_ministero_lock = threading.Lock()
_sardi_lock = threading.Lock()
_nara_lock = threading.Lock()


@app.on_event("startup")
def startup():
    init_db()
    init_usage_table()
    rti._init_tables()


@app.get("/", response_class=HTMLResponse)
def index():
    html_path = Path(__file__).parent / "templates" / "index.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


@app.get("/api/status")
def status():
    letters = []
    for l in IMI_PDFS:
        prog = None
        for p in get_all_progress():
            if p["lettera"] == l:
                prog = p
                break
        letters.append({
            "letter": l,
            "downloaded": is_downloaded(l),
            "progress": prog,
        })
    return {
        "letters": letters,
        "total_internati": count_internati(),
        "needs_review": count_needs_review(),
        "running": _running_letter,
    }


@app.post("/api/download/{letter}")
def api_download(letter: str):
    if letter not in IMI_PDFS:
        raise HTTPException(status_code=400, detail="Lettera non valida")
    try:
        download_letter(letter)
        return {"ok": True, "letter": letter}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/download-all")
def api_download_all():
    results = []
    for l in IMI_PDFS:
        if is_downloaded(l):
            results.append({"letter": l, "status": "already_downloaded"})
            continue
        try:
            download_letter(l)
            results.append({"letter": l, "status": "ok"})
        except Exception as e:
            results.append({"letter": l, "status": "error", "error": str(e)})
    return {"results": results}


# ─── Fondi archivistici (Ufficio Storico SME) ───

@app.get("/api/fondi")
def api_fondi_list():
    return {
        "fondi": get_fondi_summary(),
        "count_fondi": count_fondi(),
        "count_menzioni": count_menzioni(),
        "running": _running_fondo,
    }


@app.get("/api/fondi/available")
def api_fondi_available():
    try:
        urls = fondi.list_pdf_urls()
        return {"urls": urls, "count": len(urls)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/fondi/download-all")
def api_fondi_download_all():
    if _running_fondo:
        raise HTTPException(status_code=409, detail="Estrazione fondi in corso")
    return {"results": fondi.download_all()}


@app.post("/api/fondi/extract-all")
def api_fondi_extract_all(parallel: int = 2, engine: str = "auto"):
    global _running_fondo
    if not _fondi_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Estrazione fondi gia in corso")
    fondi.clear_stop_request()

    def run():
        global _running_fondo
        try:
            urls = fondi.list_pdf_urls()
            for u in urls:
                if fondi.is_stop_requested():
                    break
                _running_fondo = fondi._local_name(u)
                try:
                    fondi.extract_fondo(u, resume=True, engine=engine, parallel=parallel)
                except Exception as e:
                    print(f"Errore fondo {u}: {e}")
        finally:
            _running_fondo = None
            _fondi_lock.release()

    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Estrazione fondi avviata"}


@app.post("/api/fondi/stop")
def api_fondi_stop():
    if not _fondi_lock.locked():
        raise HTTPException(status_code=400, detail="Nessuna estrazione fondi in corso")
    fondi.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


@app.delete("/api/fondi/{file_pdf}")
def api_fondi_delete(file_pdf: str):
    n = delete_fondo_file(file_pdf)
    return {"ok": True, "deleted": n}


@app.get("/api/search")
def api_search(q: str, limit: int = 100):
    if not q or len(q.strip()) < 2:
        raise HTTPException(status_code=400, detail="Termine di ricerca troppo corto")
    return search_all(q.strip(), limit=limit)


# ─── Decorati (Albi della Memoria - ISTORECO) ───

@app.get("/api/decorati")
def api_decorati_status():
    return {
        "count": count_decorati(),
        "albi": get_decorati_albi(),
        "progress": decorati.get_progress(),
    }


@app.post("/api/decorati/scrape")
def api_decorati_scrape(albo_id: str = decorati.DEFAULT_ALBO, details: bool = True):
    if not _decorati_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Scraping decorati gia in corso")
    decorati.clear_stop_request()

    def run():
        try:
            decorati.scrape_albo(albo_id, resume=True, fetch_details=details)
        except Exception as e:
            print(f"Errore scraping decorati: {e}")
        finally:
            _decorati_lock.release()

    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Scraping decorati avviato", "albo_id": albo_id}


@app.post("/api/decorati/stop")
def api_decorati_stop():
    if not _decorati_lock.locked():
        raise HTTPException(status_code=400, detail="Nessuno scraping decorati in corso")
    decorati.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


# ─── Entità e Collegamenti cross-dataset ───

@app.get("/api/entita")
def api_entita_stats():
    return {
        "count_entita": count_entita(),
        "count_collegamenti": count_collegamenti(),
        "progress": linker.get_progress(),
    }


@app.post("/api/entita/build")
def api_entita_build():
    if not _linker_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Costruzione entita gia in corso")
    linker.clear_stop_request()

    def run():
        try:
            linker.build_links()
        except Exception as e:
            print(f"Errore linker: {e}")
        finally:
            _linker_lock.release()

    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Estrazione entita e collegamenti avviata"}


@app.post("/api/entita/stop")
def api_entita_stop():
    if not _linker_lock.locked():
        raise HTTPException(status_code=400, detail="Nessuna estrazione entita in corso")
    linker.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


@app.get("/api/entita/search")
def api_entita_search(q: str, limit: int = 50):
    if not q or len(q.strip()) < 2:
        raise HTTPException(status_code=400, detail="Termine troppo corto")
    return {"results": search_entita(q.strip(), limit=limit)}


@app.get("/api/entita/{entita_id}")
def api_entita_detail(entita_id: int):
    result = get_collegamenti_entita(entita_id)
    if not result:
        raise HTTPException(status_code=404, detail="Entita non trovata")
    return result


# ─── Ricerca AI ───

@app.post("/api/ai-research")
def api_ai_research(data: dict = Body(...)):
    query = data.get("query", "").strip()
    provider = data.get("provider", "gpt")
    limit = data.get("limit", 20)
    if not query:
        raise HTTPException(status_code=400, detail="Query vuota")
    if provider not in ("gpt", "mistral", "perplexity", "claude", "all"):
        raise HTTPException(status_code=400, detail="Provider non valido. Usa: gpt, mistral, perplexity, claude, all")
    try:
        if provider == "all":
            result = ai_research.research_all(query, limit=limit)
        else:
            result = ai_research.research(query, provider=provider, limit=limit)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore ricerca AI: {str(e)}")


@app.get("/api/ai-research/history")
def api_ai_research_history(limit: int = 20):
    return {"ricerche": get_ai_ricerche(limit=limit)}


# ─── Caduti Albo d'Oro (Cimeetrincee) ───

@app.get("/api/albooro")
def api_albooro_stats():
    return {
        "count": caduti_albooro.count_caduti_albooro(),
        "volumes": caduti_albooro.get_volumes_summary(),
        "progress": caduti_albooro.get_progress(),
    }


@app.post("/api/albooro/scrape")
def api_albooro_scrape():
    if not _albooro_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Scraping Albo d'Oro gia in corso")
    caduti_albooro.clear_stop_request()
    def run():
        try:
            caduti_albooro.scrape_all(resume=True)
        except Exception as e:
            print(f"Errore albooro: {e}")
        finally:
            _albooro_lock.release()
    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Scraping Albo d'Oro avviato"}


@app.post("/api/albooro/stop")
def api_albooro_stop():
    caduti_albooro.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


# ─── Caduti Bolognesi (Museo Risorgimento BO) ───

@app.get("/api/bologna")
def api_bologna_stats():
    return {
        "count": caduti_bologna.count_caduti_bologna(),
        "progress": caduti_bologna.get_progress(),
    }


@app.post("/api/bologna/scrape")
def api_bologna_scrape():
    if not _bologna_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Scraping Bologna gia in corso")
    caduti_bologna.clear_stop_request()
    def run():
        try:
            caduti_bologna.scrape_all(resume=True)
        except Exception as e:
            print(f"Errore bologna: {e}")
        finally:
            _bologna_lock.release()
    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Scraping Caduti Bolognesi avviato"}


@app.post("/api/bologna/stop")
def api_bologna_stop():
    caduti_bologna.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


# ─── NARA Catalog (After Action Reports WW2) ───

_nara_catalog_lock = threading.Lock()


@app.get("/api/nara_catalog")
def api_nara_catalog_stats():
    return {
        "count": nara_catalog.count_documenti_nara_catalog(),
        "progress": nara_catalog.get_progress(),
    }


@app.post("/api/nara_catalog/scrape")
def api_nara_catalog_scrape():
    if not _nara_catalog_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Scraping NARA Catalog già in corso")
    nara_catalog.clear_stop_request()
    def run():
        try:
            nara_catalog.scrape_all(resume=True)
        except Exception as e:
            print(f"Errore nara_catalog: {e}")
        finally:
            _nara_catalog_lock.release()
    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Scraping NARA Catalog avviato"}


@app.post("/api/nara_catalog/stop")
def api_nara_catalog_stop():
    nara_catalog.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


# ─── CWGC (Commonwealth War Graves Commission) ───

@app.get("/api/cwgc")
def api_cwgc_stats():
    return {
        "count": caduti_cwgc.count_caduti_cwgc(),
        "progress": caduti_cwgc.get_progress(),
    }


@app.post("/api/cwgc/scrape")
def api_cwgc_scrape():
    if not _cwgc_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Scraping CWGC gia in corso")
    caduti_cwgc.clear_stop_request()
    def run():
        try:
            caduti_cwgc.scrape_all(resume=True)
        except Exception as e:
            print(f"Errore cwgc: {e}")
        finally:
            _cwgc_lock.release()
    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Scraping CWGC avviato"}


@app.post("/api/cwgc/stop")
def api_cwgc_stop():
    caduti_cwgc.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


# ─── Caduti Francia WW1 (Mémoire des Hommes) ───

@app.get("/api/francia_ww1")
def api_francia_ww1_stats():
    return {
        "count": caduti_francia_ww1.count_caduti_francia_ww1(),
        "progress": caduti_francia_ww1.get_progress(),
    }


# ─── Decorati Nastro Azzurro ───

@app.get("/api/nastroazzurro")
def api_nastroazzurro_stats():
    return {
        "count": decorati_nastroazzurro.count_decorati_na(),
        "progress": decorati_nastroazzurro.get_progress(),
    }


# ─── Ministero Difesa - Caduti 1a GM ───

@app.get("/api/ministero")
def api_ministero_stats():
    return {
        "count": caduti_ministero.count_caduti_ministero(),
        "progress": caduti_ministero.get_progress(),
    }


@app.post("/api/ministero/scrape")
def api_ministero_scrape():
    if not _ministero_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Scraping Ministero gia in corso")
    caduti_ministero.clear_stop_request()
    def run():
        try:
            caduti_ministero.scrape_all(resume=True)
        except Exception as e:
            print(f"Errore ministero: {e}")
        finally:
            _ministero_lock.release()
    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Scraping Ministero Difesa avviato"}


@app.post("/api/ministero/stop")
def api_ministero_stop():
    caduti_ministero.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


# ─── Caduti Sardi (Unione Sarda) ───

@app.get("/api/sardi")
def api_sardi_stats():
    return {
        "count": caduti_sardi.count_caduti_sardi(),
        "progress": caduti_sardi.get_progress(),
    }


@app.post("/api/sardi/scrape")
def api_sardi_scrape():
    if not _sardi_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Scraping Sardi gia in corso")
    caduti_sardi.clear_stop_request()
    def run():
        try:
            caduti_sardi.scrape_all(resume=True)
        except Exception as e:
            print(f"Errore sardi: {e}")
        finally:
            _sardi_lock.release()
    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Scraping Caduti Sardi avviato"}


@app.post("/api/sardi/stop")
def api_sardi_stop():
    caduti_sardi.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


# ─── NARA T-315 Roll 1299 ───

@app.get("/api/nara")
def api_nara_stats():
    return {
        "count": nara_t315_ocr.count_documenti_nara(),
        "progress": nara_t315_ocr.get_progress(),
    }


@app.post("/api/nara/scrape")
def api_nara_scrape():
    if not _nara_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Scraping NARA gia in corso")
    nara_t315_ocr.clear_stop_request()
    def run():
        try:
            nara_t315_ocr.scrape_all(resume=True)
        except Exception as e:
            print(f"Errore nara: {e}")
        finally:
            _nara_lock.release()
    threading.Thread(target=run, daemon=True).start()
    return {"ok": True, "message": "Scraping NARA T315 R1299 avviato"}


@app.post("/api/nara/stop")
def api_nara_stop():
    nara_t315_ocr.request_stop()
    return {"ok": True, "message": "Richiesta arresto ricevuta"}


@app.post("/api/extract/stop")
def api_stop_extract():
    if not _extraction_lock.locked():
        raise HTTPException(status_code=400, detail="Nessuna estrazione in corso")
    request_stop()
    return {"ok": True, "message": "Richiesta arresto estrazione ricevuta"}


@app.post("/api/extract/resume")
def api_resume_extract(engine: str = "openai", parallel: int = 1):
    latest = get_latest_active_progress()
    if not latest:
        raise HTTPException(status_code=400, detail="Nessuna estrazione sospesa trovata")
    letter = latest["lettera"]
    return api_extract(letter, engine=engine, parallel=parallel)


@app.post("/api/extract/cancel")
def api_cancel_extract(data: dict = Body(...)):
    letter = data.get("letter")
    if not letter:
        raise HTTPException(status_code=400, detail="Letter parameter richiesto")
    deleted = delete_letter(letter)
    request_stop()
    clear_last_completed_letter()
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Lettera non trovata")
    return {"ok": True, "deleted": deleted}


@app.get("/api/extract/status")
def api_extract_status():
    latest = get_latest_active_progress()
    status = latest["status"] if latest else "idle"
    processed = latest["processed_pages"] if latest else 0
    total = latest["total_pages"] if latest else 0
    return {"status": status, "processed": processed, "total": total, "last_completed": get_last_completed_letter()}


@app.post("/api/extract/{letter}")
def api_extract(letter: str, engine: str = "openai", parallel: int = 1):
    global _running_letter
    if letter not in IMI_PDFS:
        raise HTTPException(status_code=400, detail="Lettera non valida")
    if not is_downloaded(letter):
        raise HTTPException(status_code=400, detail=f"PDF lettera {letter} non scaricato. Scaricalo prima.")
    if not _extraction_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Estrazione gia in corso. Attendi il completamento.")
    clear_stop_request()

    def run():
        global _running_letter
        _running_letter = letter
        try:
            extract_letter(letter, resume=True, engine=engine, parallel=parallel)
        except Exception as e:
            print(f"Errore estrazione {letter}: {e}")
        finally:
            _running_letter = None
            _extraction_lock.release()

    t = threading.Thread(target=run, daemon=True)
    t.start()
    return {"ok": True, "letter": letter, "message": f"Estrazione avviata (engine: {engine}, parallel: {parallel})"}


@app.post("/api/extract-all")
def api_extract_all(engine: str = "openai", parallel: int = 1):
    global _running_letter
    if not _extraction_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Estrazione gia in corso")
    clear_stop_request()

    downloaded = get_downloaded_letters()
    if not downloaded:
        raise HTTPException(status_code=400, detail="Nessun PDF scaricato. Scarica prima i PDF.")

    def run_all():
        global _running_letter
        for letter in downloaded:
            _running_letter = letter
            try:
                extract_letter(letter, resume=True, engine=engine, parallel=parallel)
            except Exception as e:
                print(f"Errore estrazione {letter}: {e}")
            if is_stop_requested():
                break
        _running_letter = None
        _extraction_lock.release()

    t = threading.Thread(target=run_all, daemon=True)
    t.start()
    return {"ok": True, "message": f"Estrazione batch avviata per {len(downloaded)} lettere (engine: {engine}, parallel: {parallel})"}


@app.get("/api/internati")
def api_internati(limit: int = 100, offset: int = 0, lettera: str = None, needs_review: bool = False):
    rows = get_internati(limit=limit, offset=offset, lettera=lettera, needs_review_only=needs_review)
    total = count_needs_review() if needs_review else count_internati()
    return {"internati": rows, "total": total, "limit": limit, "offset": offset}


@app.put("/api/internati/{rid}")
def api_update_internato(rid: int, data: dict = Body(...)):
    if not get_internato_by_id(rid):
        raise HTTPException(status_code=404, detail="Internato non trovato")
    updated = update_internato(rid, data)
    if not updated:
        raise HTTPException(status_code=400, detail="Nessun campo valido da aggiornare")
    return {"ok": True, "id": rid, "data": get_internato_by_id(rid)}


@app.delete("/api/internati/{rid}")
def api_delete_internato(rid: int):
    if not delete_internato(rid):
        raise HTTPException(status_code=404, detail="Internato non trovato")
    return {"ok": True, "deleted": rid}


@app.get("/api/internati/{rid}")
def api_get_internato(rid: int):
    row = get_internato_by_id(rid)
    if not row:
        raise HTTPException(status_code=404, detail="Internato non trovato")
    return row


@app.post("/api/validate-locations/{rid}")
def api_validate_locations(rid: int):
    row = get_internato_by_id(rid)
    if not row:
        raise HTTPException(status_code=404, detail="Internato non trovato")
    record = dict(row)
    record = validate_record_locations(record)
    update_data = {}
    if "needs_review" in record:
        update_data["needs_review"] = record["needs_review"]
    if "review_reason" in record:
        update_data["review_reason"] = record["review_reason"]
    if "luogo_validato" in record:
        update_data["luogo_validato"] = record["luogo_validato"]
    if update_data:
        update_internato(rid, update_data)
    return {"ok": True, "id": rid, "needs_review": record.get("needs_review", False), "review_reason": record.get("review_reason", ""), "luogo_validato": True}


@app.post("/api/validate-all-locations")
def api_validate_all_locations(limit: int = 500):
    """Validate locations for all records that haven't been validated yet."""
    from database import get_conn
    conn = get_conn()
    rows = conn.execute("SELECT * FROM internati WHERE luogo_validato = 0 LIMIT ?", (limit,)).fetchall()
    conn.close()
    validated = 0
    flagged = 0
    for row in rows:
        record = dict(row)
        record = validate_record_locations(record)
        update_data = {"luogo_validato": True}
        if record.get("needs_review"):
            update_data["needs_review"] = True
            update_data["review_reason"] = record.get("review_reason", "")
            flagged += 1
        update_internato(record["id"], update_data)
        validated += 1
    return {"ok": True, "validated": validated, "flagged": flagged}


@app.get("/api/credits")
def api_credits():
    return get_usage_summary()


@app.get("/api/export/excel")
def api_export_excel():
    path = export_excel()
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="export_internati.xlsx",
    )


@app.get("/api/export/csv")
def api_export_csv():
    path = export_csv()
    return FileResponse(
        path,
        media_type="text/csv",
        filename="export_internati.csv",
    )


# ─── Web Scraping ───

@app.post("/api/scrape")
def api_scrape(url: str = Body(..., embed=True), max_depth: int = Body(2, embed=True)):
    if not url.startswith("http"):
        raise HTTPException(status_code=400, detail="URL non valido. Deve iniziare con http:// o https://")
    started = run_scrape(url, max_depth=max_depth)
    if not started:
        raise HTTPException(status_code=409, detail="Scraping gia in corso")
    return {"ok": True, "message": f"Scraping avviato per {url}"}


@app.get("/api/scrape/status")
def api_scrape_status():
    return get_scrape_status()


# ─── File Upload ───

@app.post("/api/upload")
async def api_upload(file: UploadFile = File(...), engine: str = Form("openai")):
    ftype = detect_file_type(file.filename)
    if ftype == "unknown":
        raise HTTPException(status_code=400, detail=f"Formato non supportato. Usare CSV, Excel, PDF o immagini (jpg/png/tiff)")
    local_path = UPLOAD_DIR / file.filename
    with open(local_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    try:
        if ftype == "csv":
            result = import_csv(local_path)
        elif ftype == "excel":
            result = import_excel(local_path)
        elif ftype == "pdf":
            result = import_pdf(local_path, engine=engine)
        elif ftype == "image":
            result = import_image(local_path)
        else:
            raise HTTPException(status_code=400, detail="Tipo file non supportato")
        return {"ok": True, "file_type": ftype, "result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore importazione: {str(e)}")


@app.post("/api/upload/preview")
async def api_upload_preview(file: UploadFile = File(...)):
    """Upload a file and return detected headers + suggested mapping without importing."""
    ftype = detect_file_type(file.filename)
    if ftype not in ("csv", "excel"):
        raise HTTPException(status_code=400, detail="Preview disponibile solo per CSV e Excel")
    local_path = UPLOAD_DIR / f"preview_{file.filename}"
    with open(local_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    try:
        if ftype == "csv":
            import csv as csvmod
            with open(local_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csvmod.DictReader(f)
                headers = reader.fieldnames or []
                sample = [dict(next(reader)) for _ in range(3)]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(str(local_path), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])] if rows else []
            sample = [dict(zip(headers, r)) for r in rows[1:4]] if len(rows) > 1 else []
        suggested = _suggest_mapping(headers)
        return {"ok": True, "file_type": ftype, "headers": list(headers), "sample": sample, "suggested_mapping": suggested, "db_columns": COLUMNS}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore preview: {str(e)}")


@app.post("/api/upload/import")
async def api_upload_import(
    file: UploadFile = File(...),
    mapping: str = Form(...),
    lettera: str = Form("IMPORT"),
):
    """Import a CSV/Excel file with a user-defined column mapping."""
    ftype = detect_file_type(file.filename)
    if ftype not in ("csv", "excel"):
        raise HTTPException(status_code=400, detail="Import con mapping disponibile solo per CSV e Excel")
    local_path = UPLOAD_DIR / file.filename
    with open(local_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    try:
        column_mapping = json.loads(mapping)
        if ftype == "csv":
            result = import_csv(local_path, column_mapping=column_mapping, lettera=lettera)
        else:
            result = import_excel(local_path, column_mapping=column_mapping, lettera=lettera)
        return {"ok": True, "file_type": ftype, "result": result}
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Mapping JSON non valido")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore importazione: {str(e)}")


# ─── Archivio Fonti Documentali ───

archivio_fonti._init_table()


@app.get("/api/archivio")
def api_archivio_stats():
    return {
        "count": archivio_fonti._count_saved(),
        "storage_dir": str(archivio_fonti.STORAGE_DIR),
    }


@app.post("/api/archivio/query")
def api_archivio_query(body: dict = Body(...)):
    return archivio_fonti.query_archivio(
        unita=body.get("unita", ""),
        teatro=body.get("teatro", ""),
        data_da=body.get("data_da", ""),
        data_a=body.get("data_a", ""),
        tipo_documento=body.get("tipo_documento", ""),
        archivio=body.get("archivio", ""),
        fondo=body.get("fondo", ""),
        conflitto=body.get("conflitto", ""),
        testo_libero=body.get("testo_libero", ""),
        solo_leggibili=body.get("solo_leggibili", False),
        limit=body.get("limit", 50),
        offset=body.get("offset", 0),
    )


@app.get("/api/archivio/file/{sha256}")
def api_archivio_file(sha256: str):
    p = archivio_fonti.get_file_path(sha256)
    if not p or not p.exists():
        raise HTTPException(status_code=404, detail="File non trovato")
    media = {
        ".pdf": "application/pdf",
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png", ".tiff": "image/tiff", ".tif": "image/tiff",
    }.get(p.suffix.lower(), "application/octet-stream")
    return FileResponse(str(p), media_type=media, filename=p.name)


@app.post("/api/archivio/ingest")
async def api_archivio_ingest(file: UploadFile = File(...), meta: str = Form(default="{}")):
    try:
        meta_dict = json.loads(meta)
    except Exception:
        meta_dict = {}
    tmp = archivio_fonti.STORAGE_DIR / f"_upload_{file.filename}"
    try:
        content = await file.read()
        tmp.write_bytes(content)
        result = archivio_fonti.ingest_file(
            tmp,
            archivio=meta_dict.get("archivio", ""),
            fondo=meta_dict.get("fondo", ""),
            serie=meta_dict.get("serie", ""),
            busta=meta_dict.get("busta", ""),
            fascicolo=meta_dict.get("fascicolo", ""),
            segnatura=meta_dict.get("segnatura", ""),
            titolo_documento=meta_dict.get("titolo_documento", ""),
            unita_principale=meta_dict.get("unita_principale", ""),
            livello_unita=meta_dict.get("livello_unita", ""),
            unita_superiore=meta_dict.get("unita_superiore", ""),
            teatro_operazioni=meta_dict.get("teatro_operazioni", ""),
            nazione_forza=meta_dict.get("nazione_forza", ""),
            data_inizio=meta_dict.get("data_inizio", ""),
            data_fine=meta_dict.get("data_fine", ""),
            data_raw=meta_dict.get("data_raw", ""),
            tipo_documento=meta_dict.get("tipo_documento", "corsivo_illeggibile"),
            conflitto=meta_dict.get("conflitto", ""),
            unita_citate=meta_dict.get("unita_citate", []),
            luoghi_citati=meta_dict.get("luoghi_citati", []),
            parole_chiave=meta_dict.get("parole_chiave", []),
            attendibilita_fonte=meta_dict.get("attendibilita_fonte", 3),
            note=meta_dict.get("note", ""),
            fonte_acquisizione=meta_dict.get("fonte_acquisizione", "UPLOAD_MANUALE"),
            ocr_status=meta_dict.get("ocr_status", "pending"),
        )
        return {"ok": True, "result": result}
    finally:
        if tmp.exists():
            tmp.unlink()


@app.post("/api/archivio/retrofit_nara_t315")
def api_archivio_retrofit_nara():
    threading.Thread(target=archivio_fonti.retrofit_nara_t315, daemon=True).start()
    return {"ok": True, "message": "Retrofit NARA T315 avviato in background"}


# ─── Hippocampal Memory Router ───

memory_router._init_tables()


@app.post("/api/memory/query")
def api_memory_query(body: dict = Body(...)):
    q = body.get("query", "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="query mancante")
    depth = body.get("depth", None)
    use_cloud = body.get("use_cloud_fallback", True)
    return memory_router.route_query(q, depth=depth, use_cloud_fallback=use_cloud)


@app.get("/api/memory/traces")
def api_memory_traces(limit: int = 20):
    return {"traces": memory_router.get_recent_traces(limit=limit)}


@app.get("/api/memory/consolidated")
def api_memory_consolidated(limit: int = 20):
    return {"consolidated": memory_router.get_consolidated(limit=limit)}


@app.get("/api/memory/stats")
def api_memory_stats():
    traces = memory_router.get_recent_traces(limit=100)
    avg_ms = round(sum(t["response_ms"] or 0 for t in traces) / max(len(traces), 1), 1)
    tokens_saved = sum(t["tokens_saved_estimate"] or 0 for t in traces)
    return {
        "total_traces": memory_router.count_traces(),
        "consolidated_topics": len(memory_router.get_consolidated(limit=1000)),
        "avg_response_ms": avg_ms,
        "tokens_saved_estimate": tokens_saved,
    }


# ─── Source Locator (indice leggero fonti) ───

source_locator._init_tables()


@app.post("/api/sources/index")
def api_sources_index(body: dict = Body(...)):
    """Registra la scheda di collocazione di una fonte (solo metadati)."""
    if not body.get("titolo") and not body.get("segnatura"):
        raise HTTPException(status_code=400, detail="serve almeno titolo o segnatura")
    return source_locator.register_source_metadata(**body)


@app.post("/api/sources/search")
def api_sources_search(body: dict = Body(...)):
    q = body.get("query", "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="query mancante")
    return source_locator.find_candidate_sources(q, limit=body.get("limit", 20))


@app.post("/api/sources/fetch/{source_id}")
def api_sources_fetch(source_id: int, force: bool = False, permanent: bool = False):
    result = source_locator.fetch_source_on_demand(source_id, force=force, permanent=permanent)
    if not result.get("ok") and "non trovata" in result.get("error", ""):
        raise HTTPException(status_code=404, detail=result["error"])
    return result


@app.post("/api/sources/context")
def api_sources_context(body: dict = Body(...)):
    """Contesto minimo per l'AI: metadati + excerpt solo da cache locale.
    L'AI non scarica mai direttamente: allow_fetch è gestito dal backend."""
    q = body.get("query", "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="query mancante")
    return source_locator.build_minimal_context_for_ai(
        q,
        max_sources=body.get("max_sources", 5),
        allow_fetch=bool(body.get("allow_fetch", False)),
    )


@app.get("/api/sources/stats")
def api_sources_stats():
    return source_locator.get_stats()


# ─── Source Federation Layer ───

@app.get("/api/providers")
def api_providers():
    """Lista tutti i provider registrati nel federation layer."""
    return {"providers": list_providers(), "count": len(list_providers())}


@app.get("/api/providers/{provider_name}")
def api_provider_detail(provider_name: str):
    """Dettaglio di un provider specifico."""
    p = get_provider(provider_name)
    if not p:
        raise HTTPException(status_code=404, detail=f"provider '{provider_name}' non trovato")
    return {
        "name": p.name,
        "display_name": p.display_name,
        "country": p.country,
        "archive_name": p.archive_name,
        "base_url": p.base_url,
        "authorized_domains": list(p.authorized_domains),
        "cache_ttl_days": p.cache_ttl_days,
    }


@app.post("/api/source/search")
def api_source_search(body: dict = Body(...)):
    """Ricerca federata across provider. Non scarica documenti."""
    q = body.get("query", "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="query mancante")
    cues = body.get("cues", {})
    providers = body.get("providers")
    filters = body.get("filters", {})
    results = federated_search(q, cues=cues, providers=providers, filters=filters)
    return {"query": q, "results": results, "count": len(results)}


@app.post("/api/source/fetch")
def api_source_fetch(body: dict = Body(...)):
    """Fetch on-demand di un documento. Solo backend, solo domini autorizzati."""
    source_id = body.get("source_id")
    if not source_id:
        raise HTTPException(status_code=400, detail="source_id mancante")
    return fed_fetch_source(int(source_id))


@app.get("/api/source/cache")
def api_source_cache():
    """Lista file in cache."""
    import sqlite3
    from database import DB_PATH
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM source_fetch_cache ORDER BY fetched_at DESC LIMIT 50")
    rows = [{k: r[k] for k in r.keys()} for r in cur.fetchall()]
    conn.close()
    return {"cache": rows, "count": len(rows)}


@app.get("/api/source/stats")
def api_source_stats():
    """Statistiche federation layer."""
    return get_federation_stats()


@app.post("/api/source/reindex")
def api_source_reindex(body: dict = Body(...)):
    """Re-index: ricerca metadati da provider e registra in fonti_indice."""
    q = body.get("query", "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="query mancante")
    providers = body.get("providers")
    filters = body.get("filters", {})
    results = federated_search(q, providers=providers, filters=filters)
    registered = 0
    for r in results:
        if "error" in r:
            continue
        p = get_provider(r.get("provider", ""))
        if p:
            try:
                p.register_in_db(r)
                registered += 1
            except Exception:
                pass
    return {"query": q, "found": len(results), "registered": registered}


# ─── Soldier Dashboard ───

@app.get("/api/soldiers/{soldier_id}/dashboard")
def api_soldier_dashboard(soldier_id: int):
    """Dashboard investigativa: dati soldato + timeline + fonti locali + esterne."""
    result = soldier_dashboard.get_soldier_dashboard(soldier_id)
    if not result.get("ok"):
        raise HTTPException(status_code=404, detail=result.get("error", "soldato non trovato"))
    return result


@app.get("/api/soldiers/{soldier_id}/sources")
def api_soldier_sources(soldier_id: int):
    """Solo fonti (locali + esterne) per un soldato."""
    result = soldier_dashboard.get_soldier_sources(soldier_id)
    if not result.get("ok"):
        raise HTTPException(status_code=404, detail=result.get("error", "soldato non trovato"))
    return result


# ─── Biografie (sintesi AI da fonti verificate, con fallback multi-provider) ───

@app.post("/api/biography")
def api_generate_biography(data: dict = Body(...)):
    """Genera una biografia/dossier narrativo con l'AI, usando SOLO dati e
    fonti gia' verificate (mai le fonti esterne solo candidate/da recuperare).

    Body: {subject_type: "soldier"|"event", soldier_id?, query?, provider?}
    provider e' opzionale: se assente, prova gpt -> claude -> mistral ->
    perplexity in automatico e usa il primo che risponde.
    """
    subject_type = data.get("subject_type", "soldier")
    provider = data.get("provider") or None

    if subject_type == "soldier":
        soldier_id = data.get("soldier_id")
        if not soldier_id:
            raise HTTPException(status_code=400, detail="soldier_id richiesto per subject_type=soldier")
        result = biography.generate_soldier_biography(int(soldier_id), provider=provider)
    elif subject_type == "event":
        query = (data.get("query") or "").strip()
        if not query:
            raise HTTPException(status_code=400, detail="query richiesta per subject_type=event")
        result = biography.generate_event_biography(query, provider=provider)
    else:
        raise HTTPException(status_code=400, detail="subject_type deve essere 'soldier' o 'event'")

    if result.get("error") and "risposta" not in result:
        raise HTTPException(status_code=404 if "non trovato" in result["error"] else 502, detail=result["error"])
    return result


@app.post("/api/sources/analyze")
def api_sources_analyze(body: dict = Body(...)):
    """Prepara contesto minimo per AI analysis. Backend seleziona fonti."""
    source_ids = body.get("source_ids", [])
    query = body.get("query", "")
    if not source_ids:
        raise HTTPException(status_code=400, detail="source_ids mancante")
    return soldier_dashboard.analyze_sources(source_ids, query)


# ─── Research-to-Index ───

import research_to_index as rti


@app.post("/api/research/query")
def api_research_query(body: dict = Body(...)):
    """Auto-index: cerca locale → se non trova, crea soggetto + arricchisce con fonti esterne."""
    query = body.get("query", "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="query mancante")
    try:
        result = rti.auto_index_if_not_found(query)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore auto-index: {str(e)}")


@app.post("/api/research/auto-index")
def api_research_auto_index(body: dict = Body(...)):
    """Forza creazione soggetto di ricerca (anche se esiste gia' nel DB locale)."""
    query = body.get("query", "").strip()
    subject_type = body.get("subject_type")
    if not query:
        raise HTTPException(status_code=400, detail="query mancante")
    try:
        subject = rti.create_minimal_subject_from_query(query, subject_type)
        subject_id = subject["id"]
        if not subject.get("already_existed"):
            cues = {"persona": query}
            fed_results = rti.federated_search(query, cues=cues)
            indexed = []
            for result in fed_results:
                if "error" in result:
                    continue
                source_id = rti.upsert_source_locator(result)
                if source_id:
                    conf = result.get("score", result.get("confidence", 0.3))
                    rti.link_subject_to_source(subject_id, source_id, "mentions", conf,
                                               f"Provider: {result.get('provider', '?')}")
                    indexed.append({
                        "source_id": source_id,
                        "archivio": result.get("archivio"),
                        "titolo": (result.get("titolo") or "")[:80],
                        "provider": result.get("provider"),
                        "score": result.get("score", 0),
                    })
            rti.update_subject_confidence(subject_id)
            rti.identify_research_gaps(subject_id)
            subject["indexed_sources"] = indexed
            subject["sources_count"] = len(indexed)
        return subject
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore creazione soggetto: {str(e)}")


@app.get("/api/research/subjects")
def api_research_subjects(
    subject_type: str = None,
    status: str = None,
    min_confidence: float = None,
    limit: int = 50,
    offset: int = 0,
):
    """Lista soggetti di ricerca con filtri opzionali."""
    from database import get_conn
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    sql = "SELECT * FROM research_subjects WHERE 1=1"
    params = []
    if subject_type:
        sql += " AND subject_type = ?"
        params.append(subject_type)
    if status:
        sql += " AND status = ?"
        params.append(status)
    if min_confidence is not None:
        sql += " AND confidence >= ?"
        params.append(min_confidence)
    sql += " ORDER BY updated_at DESC LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    rows = conn.execute(sql, params).fetchall()
    total = conn.execute(
        "SELECT COUNT(*) FROM research_subjects" +
        (" WHERE subject_type = ?" if subject_type else "") +
        (" AND status = ?" if status else ""),
        [p for p in params[:-2] if p is not None]
    ).fetchone()[0] if (subject_type or status) else conn.execute(
        "SELECT COUNT(*) FROM research_subjects"
    ).fetchone()[0]
    conn.close()
    return {
        "subjects": [dict(r) for r in rows],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


@app.get("/api/research/subjects/{subject_id}")
def api_research_subject_detail(subject_id: int):
    """Dettaglio soggetto con fonti collegate."""
    from database import get_conn
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    subject = conn.execute(
        "SELECT * FROM research_subjects WHERE id = ?", (subject_id,)
    ).fetchone()
    if not subject:
        conn.close()
        raise HTTPException(status_code=404, detail="Soggetto non trovato")
    sources = conn.execute(
        """SELECT rs.*, fi.archivio, fi.titolo, fi.url_catalogo, fi.access_type,
                  fi.confidence as src_confidence, fi.segnatura
           FROM research_subject_sources rs
           JOIN fonti_indice fi ON rs.source_locator_id = fi.id
           WHERE rs.subject_id = ?
           ORDER BY rs.confidence DESC""",
        (subject_id,)
    ).fetchall()
    gaps = conn.execute(
        "SELECT * FROM research_gaps WHERE subject_id = ? ORDER BY priority",
        (subject_id,)
    ).fetchall()
    conn.close()
    return {
        "subject": dict(subject),
        "sources": [dict(s) for s in sources],
        "gaps": [dict(g) for g in gaps],
    }


@app.get("/api/research/subjects/{subject_id}/dashboard")
def api_research_subject_dashboard(subject_id: int):
    """Dashboard completa: soggetto + fonti + gaps + arricchimento."""
    from database import get_conn
    try:
        enriched = rti.enrich_subject_from_sources(subject_id)
    except Exception:
        enriched = {}
    detail = api_research_subject_detail(subject_id)
    stats = rti.get_research_stats()
    return {
        **detail,
        "enrichment": enriched,
        "stats": stats,
    }


@app.patch("/api/research/subjects/{subject_id}")
def api_research_subject_update(subject_id: int, body: dict = Body(...)):
    """Aggiorna status/confidence/campi di un soggetto."""
    from database import get_conn
    conn = get_conn()
    subject = conn.execute(
        "SELECT id FROM research_subjects WHERE id = ?", (subject_id,)
    ).fetchone()
    if not subject:
        conn.close()
        raise HTTPException(status_code=404, detail="Soggetto non trovato")
    allowed_fields = {"status", "confidence", "name", "date_start", "date_end",
                      "place", "unit", "linked_soldier_id"}
    updates = []
    params = []
    for key, value in body.items():
        if key in allowed_fields:
            updates.append(f"{key} = ?")
            params.append(value)
    if not updates:
        conn.close()
        raise HTTPException(status_code=400, detail="Nessun campo valido da aggiornare")
    updates.append("updated_at = ?")
    params.append(datetime.now().isoformat())
    params.append(subject_id)
    conn.execute(
        f"UPDATE research_subjects SET {', '.join(updates)} WHERE id = ?",
        params
    )
    conn.commit()
    conn.close()
    return {"ok": True, "subject_id": subject_id, "updated_fields": list(body.keys())}


@app.get("/api/research/gaps")
def api_research_gaps(
    status: str = "open",
    subject_id: int = None,
    limit: int = 100,
):
    """Lista gaps aperti con suggerimenti provider."""
    from database import get_conn
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    sql = """SELECT rg.*, rs.name as subject_name, rs.subject_type
             FROM research_gaps rg
             JOIN research_subjects rs ON rg.subject_id = rs.id
             WHERE rg.status = ?"""
    params = [status]
    if subject_id is not None:
        sql += " AND rg.subject_id = ?"
        params.append(subject_id)
    sql += " ORDER BY rg.priority DESC, rg.created_at DESC LIMIT ?"
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return {
        "gaps": [dict(r) for r in rows],
        "count": len(rows),
        "status_filter": status,
    }


@app.get("/api/research/stats")
def api_research_stats():
    """Statistiche Research-to-Index."""
    try:
        return rti.get_research_stats()
    except Exception as e:
        return {"initialized": False, "error": str(e)}

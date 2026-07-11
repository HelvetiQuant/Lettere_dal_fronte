import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DB_PATH = Path(__file__).parent / "ocr_lettere.db"
UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)


def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS lettere (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            file_path TEXT NOT NULL,
            mittente TEXT,
            destinatario TEXT,
            data_lettera TEXT,
            luogo TEXT,
            oggetto TEXT,
            corpo_testo TEXT,
            note TEXT,
            confidenza REAL,
            lingua TEXT,
            raw_response TEXT,
            elaborato_il TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()


def save_letter(data: dict) -> int:
    conn = get_connection()
    cur = conn.execute(
        """INSERT INTO lettere
           (filename, file_path, mittente, destinatario, data_lettera, luogo,
            oggetto, corpo_testo, note, confidenza, lingua, raw_response, elaborato_il)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            data.get("_source_file", ""),
            data.get("_file_path", ""),
            data.get("mittente"),
            data.get("destinatario"),
            data.get("data_lettera"),
            data.get("luogo"),
            data.get("oggetto"),
            data.get("corpo_testo"),
            data.get("note"),
            float(data.get("confidenza") or 0),
            data.get("lingua"),
            data.get("_raw_response", ""),
            datetime.now().isoformat(),
        ),
    )
    conn.commit()
    rid = cur.lastrowid
    conn.close()
    return rid


def get_all_letters() -> list[dict]:
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM lettere ORDER BY elaborato_il DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_letter(rid: int) -> Optional[dict]:
    conn = get_connection()
    row = conn.execute("SELECT * FROM lettere WHERE id = ?", (rid,)).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_letter(rid: int) -> bool:
    conn = get_connection()
    cur = conn.execute("DELETE FROM lettere WHERE id = ?", (rid,))
    conn.commit()
    deleted = cur.rowcount > 0
    conn.close()
    return deleted


def export_excel(output_path: str = None) -> str:
    if output_path is None:
        output_path = str(Path(__file__).parent / "export_lettere.xlsx")

    letters = get_all_letters()
    wb = Workbook()
    ws = wb.active
    ws.title = "Lettere dal Fronte"

    headers = [
        "ID", "File", "Mittente", "Destinatario", "Data Lettera",
        "Luogo", "Oggetto", "Corpo Testo", "Note", "Confidenza",
        "Lingua", "Elaborato il",
    ]

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, letter in enumerate(letters, 2):
        values = [
            letter["id"], letter["filename"], letter["mittente"] or "",
            letter["destinatario"] or "", letter["data_lettera"] or "",
            letter["luogo"] or "", letter["oggetto"] or "",
            letter["corpo_testo"] or "", letter["note"] or "",
            letter["confidenza"] or 0, letter["lingua"] or "",
            letter["elaborato_il"] or "",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 60
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 22

    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path

import json
import os
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import COLUMNS

DB_PATH = Path(__file__).parent / "imi_internati.db"


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-65536")     # 64 MB cache pagine
    conn.execute("PRAGMA temp_store=MEMORY")     # tabelle temporanee in RAM
    conn.execute("PRAGMA mmap_size=268435456")   # 256 MB memory-mapped I/O
    return conn


def init_db():
    conn = get_conn()
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS internati (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lettera TEXT NOT NULL,
            file_pdf TEXT NOT NULL,
            pagina INTEGER NOT NULL,
            {", ".join(COLUMNS)} TEXT,
            raw_text TEXT,
            elaborato_il TEXT NOT NULL,
            needs_review INTEGER DEFAULT 0,
            review_reason TEXT,
            luogo_validato INTEGER DEFAULT 0
        )
    """)
    # Add columns to existing tables if they don't exist
    for col, coltype in [("needs_review", "INTEGER DEFAULT 0"), ("review_reason", "TEXT"), ("luogo_validato", "INTEGER DEFAULT 0")]:
        try:
            conn.execute(f"ALTER TABLE internati ADD COLUMN {col} {coltype}")
        except sqlite3.OperationalError:
            pass
    conn.execute("""
        CREATE TABLE IF NOT EXISTS progress (
            lettera TEXT PRIMARY KEY,
            total_pages INTEGER,
            processed_pages INTEGER DEFAULT 0,
            status TEXT DEFAULT 'pending',
            started_at TEXT,
            finished_at TEXT
        )
    """)
    # ─── Fondi archivistici (Ufficio Storico SME) ───
    conn.execute("""
        CREATE TABLE IF NOT EXISTS fondi_archivistici (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codice_fondo TEXT,
            titolo TEXT,
            file_pdf TEXT NOT NULL,
            url TEXT,
            pagina INTEGER NOT NULL,
            descrizione TEXT,
            periodo TEXT,
            busta TEXT,
            fascicolo TEXT,
            luoghi TEXT,
            raw_text TEXT,
            elaborato_il TEXT NOT NULL
        )
    """)
    # ─── Menzioni di persone/luoghi estratte dai fondi (per ricerca incrociata) ───
    conn.execute("""
        CREATE TABLE IF NOT EXISTS menzioni (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fondo_id INTEGER,
            file_pdf TEXT NOT NULL,
            pagina INTEGER,
            tipo TEXT NOT NULL,
            cognome TEXT,
            nome TEXT,
            testo_originale TEXT,
            grado TEXT,
            reparto TEXT,
            luogo TEXT,
            data TEXT,
            contesto TEXT,
            elaborato_il TEXT NOT NULL,
            FOREIGN KEY (fondo_id) REFERENCES fondi_archivistici(id)
        )
    """)
    # ─── Decorati / Caduti (Albi della Memoria - ISTORECO) ───
    conn.execute("""
        CREATE TABLE IF NOT EXISTS decorati (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_id TEXT UNIQUE,
            albo_id TEXT,
            albo_nome TEXT,
            cognome TEXT,
            nome TEXT,
            comune_nascita TEXT,
            comune_residenza TEXT,
            data_nascita TEXT,
            data_morte TEXT,
            anno_nascita INTEGER,
            anno_morte INTEGER,
            guerra TEXT,
            grado TEXT,
            corpo_militare TEXT,
            reparto TEXT,
            decorazione TEXT,
            motivazione TEXT,
            causa_morte TEXT,
            luogo_morte TEXT,
            luogo_cattura TEXT,
            luogo_internamento TEXT,
            matricola TEXT,
            professione TEXT,
            note TEXT,
            url_scheda TEXT,
            foto_urls TEXT,
            raw_json TEXT,
            elaborato_il TEXT NOT NULL
        )
    """)
    # ─── Entità estratte da tutti i dataset (persone, luoghi, eventi) ───
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entita (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL,
            valore TEXT NOT NULL,
            valore_normalizzato TEXT,
            cognome TEXT,
            nome TEXT,
            data TEXT,
            luogo TEXT,
            contesto TEXT,
            fonte_tabella TEXT,
            fonte_id INTEGER,
            elaborato_il TEXT NOT NULL
        )
    """)
    # ─── Collegamenti tra entità e record dei vari dataset ───
    conn.execute("""
        CREATE TABLE IF NOT EXISTS collegamenti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entita_id INTEGER NOT NULL,
            tabella_origine TEXT NOT NULL,
            record_id INTEGER NOT NULL,
            tipo_collegamento TEXT,
            confidenza REAL DEFAULT 1.0,
            elaborato_il TEXT NOT NULL,
            FOREIGN KEY (entita_id) REFERENCES entita(id)
        )
    """)
    # ─── Log delle ricerche AI effettuate ───
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ai_ricerche (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            query TEXT NOT NULL,
            provider TEXT NOT NULL,
            model TEXT NOT NULL,
            risposta TEXT,
            contesto_dati TEXT,
            cost_usd REAL DEFAULT 0.0,
            elaborato_il TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_menzioni_cognome ON menzioni(cognome)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_menzioni_luogo ON menzioni(luogo)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_internati_cognome ON internati(cognome)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_decorati_cognome ON decorati(cognome)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_entita_valore ON entita(valore_normalizzato)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_entita_tipo ON entita(tipo)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_collegamenti_entita ON collegamenti(entita_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_collegamenti_record ON collegamenti(tabella_origine, record_id)")
    conn.commit()
    conn.close()


def save_decorato(data: dict) -> bool:
    conn = get_conn()
    try:
        conn.execute(
            """INSERT OR REPLACE INTO decorati
               (source_id, albo_id, albo_nome, cognome, nome, comune_nascita, comune_residenza,
                data_nascita, data_morte, anno_nascita, anno_morte, guerra, grado, corpo_militare,
                reparto, decorazione, motivazione, causa_morte, luogo_morte, luogo_cattura,
                luogo_internamento, matricola, professione, note, url_scheda, foto_urls, raw_json, elaborato_il)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (data.get("source_id"), data.get("albo_id"), data.get("albo_nome"),
             data.get("cognome"), data.get("nome"), data.get("comune_nascita"), data.get("comune_residenza"),
             data.get("data_nascita"), data.get("data_morte"), data.get("anno_nascita"), data.get("anno_morte"),
             data.get("guerra"), data.get("grado"), data.get("corpo_militare"), data.get("reparto"),
             data.get("decorazione"), data.get("motivazione"), data.get("causa_morte"), data.get("luogo_morte"),
             data.get("luogo_cattura"), data.get("luogo_internamento"), data.get("matricola"),
             data.get("professione"), data.get("note"), data.get("url_scheda"), data.get("foto_urls"),
             data.get("raw_json"), datetime.now().isoformat()),
        )
        conn.commit()
        return True
    finally:
        conn.close()


def count_decorati() -> int:
    conn = get_conn()
    row = conn.execute("SELECT COUNT(*) as c FROM decorati").fetchone()
    conn.close()
    return row["c"]


def decorato_exists(source_id: str) -> bool:
    conn = get_conn()
    row = conn.execute("SELECT 1 FROM decorati WHERE source_id = ? LIMIT 1", (source_id,)).fetchone()
    conn.close()
    return row is not None


def get_decorati_albi() -> list[dict]:
    conn = get_conn()
    rows = conn.execute(
        "SELECT albo_id, MAX(albo_nome) as albo_nome, COUNT(*) as n FROM decorati GROUP BY albo_id"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def save_fondo(codice_fondo: str, titolo: str, file_pdf: str, url: str, pagina: int,
               data: dict, raw_text: str = "") -> int:
    conn = get_conn()
    cur = conn.execute(
        """INSERT INTO fondi_archivistici
           (codice_fondo, titolo, file_pdf, url, pagina, descrizione, periodo, busta, fascicolo, luoghi, raw_text, elaborato_il)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (codice_fondo, titolo, file_pdf, url, pagina,
         data.get("descrizione"), data.get("periodo"), data.get("busta"),
         data.get("fascicolo"), data.get("luoghi"), raw_text, datetime.now().isoformat()),
    )
    conn.commit()
    fid = cur.lastrowid
    conn.close()
    return fid


STANDARD_MENZIONI_FIELDS = {
    "fondo_id", "file_pdf", "pagina", "tipo", "cognome", "nome",
    "testo_originale", "grado", "reparto", "luogo", "data", "contesto",
    "elaborato_il",
}
FIELD_FREQUENCY_THRESHOLD = 0.05  # 5%


def _ensure_field_stats_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS _menzioni_field_stats (
            field_name TEXT PRIMARY KEY,
            count INTEGER DEFAULT 0,
            added_column INTEGER DEFAULT 0
        )
    """)


def _get_menzioni_columns(conn):
    cols = conn.execute("PRAGMA table_info(menzioni)").fetchall()
    return {c["name"] for c in cols}


def _maybe_add_column(conn, field_name: str, total_menzioni: int):
    """Se il campo supera il 5% delle menzioni totali, aggiungi la colonna."""
    row = conn.execute(
        "SELECT count, added_column FROM _menzioni_field_stats WHERE field_name=?",
        (field_name,)
    ).fetchone()
    if not row or row["added_column"]:
        return False
    if total_menzioni < 100:
        return False
    if row["count"] / total_menzioni >= FIELD_FREQUENCY_THRESHOLD:
        col_name = field_name.replace(" ", "_").replace("-", "_").lower()
        existing = _get_menzioni_columns(conn)
        if col_name not in existing:
            conn.execute(f"ALTER TABLE menzioni ADD COLUMN {col_name} TEXT")
            print(f"  [SCHEMA] Aggiunta colonna '{col_name}' alla tabella menzioni "
                  f"(frequenza: {row['count']}/{total_menzioni} = {row['count']/total_menzioni:.1%})")
        conn.execute(
            "UPDATE _menzioni_field_stats SET added_column=1 WHERE field_name=?",
            (field_name,)
        )
        return True
    return False


def save_menzione(fondo_id: Optional[int], file_pdf: str, pagina: int, data: dict):
    conn = get_conn()
    _ensure_field_stats_table(conn)

    # Rileva campi non standard restituiti da GPT
    extra_fields = {k: v for k, v in data.items()
                    if k not in STANDARD_MENZIONI_FIELDS and v is not None and str(v).strip()}
    if extra_fields:
        total = conn.execute("SELECT COUNT(*) as c FROM menzioni").fetchone()["c"]
        for fname in extra_fields:
            conn.execute("""
                INSERT INTO _menzioni_field_stats (field_name, count, added_column)
                VALUES (?, 1, 0)
                ON CONFLICT(field_name) SET count = count + 1
            """, (fname,))
            _maybe_add_column(conn, fname, total + 1)

    # Costruisci INSERT dinamico: include colonne standard + eventuali nuove colonne
    existing_cols = _get_menzioni_columns(conn)
    standard_values = {
        "fondo_id": fondo_id,
        "file_pdf": file_pdf,
        "pagina": pagina,
        "tipo": data.get("tipo", "persona"),
        "cognome": data.get("cognome"),
        "nome": data.get("nome"),
        "testo_originale": data.get("testo_originale"),
        "grado": data.get("grado"),
        "reparto": data.get("reparto"),
        "luogo": data.get("luogo"),
        "data": data.get("data"),
        "contesto": data.get("contesto"),
        "elaborato_il": datetime.now().isoformat(),
    }

    # Aggiungi valori per colonne extra già presenti nello schema
    for fname, fval in extra_fields.items():
        col_name = fname.replace(" ", "_").replace("-", "_").lower()
        if col_name in existing_cols:
            standard_values[col_name] = fval

    cols = list(standard_values.keys())
    placeholders = ", ".join("?" for _ in cols)
    col_list = ", ".join(cols)
    conn.execute(
        f"INSERT INTO menzioni ({col_list}) VALUES ({placeholders})",
        [standard_values[c] for c in cols],
    )
    conn.commit()
    conn.close()


def count_fondi() -> int:
    conn = get_conn()
    row = conn.execute("SELECT COUNT(DISTINCT file_pdf) as c FROM fondi_archivistici").fetchone()
    conn.close()
    return row["c"]


def count_menzioni() -> int:
    conn = get_conn()
    row = conn.execute("SELECT COUNT(*) as c FROM menzioni").fetchone()
    conn.close()
    return row["c"]


def is_fondo_page_processed(file_pdf: str, pagina: int) -> bool:
    conn = get_conn()
    row = conn.execute(
        "SELECT 1 FROM fondi_archivistici WHERE file_pdf = ? AND pagina = ? LIMIT 1",
        (file_pdf, pagina),
    ).fetchone()
    conn.close()
    return row is not None


def delete_fondo_file(file_pdf: str) -> int:
    conn = get_conn()
    cur = conn.execute("DELETE FROM fondi_archivistici WHERE file_pdf = ?", (file_pdf,))
    conn.execute("DELETE FROM menzioni WHERE file_pdf = ?", (file_pdf,))
    conn.execute("DELETE FROM progress WHERE lettera = ?", (f"FONDO:{file_pdf}",))
    conn.commit()
    n = cur.rowcount
    conn.close()
    return n


def _tokenize(term: str) -> list[str]:
    """Divide la query in token utili (almeno 2 caratteri)."""
    import re
    tokens = re.findall(r"[A-Za-z0-9àèéìòù']{2,}", term.lower())
    return [t for t in tokens if len(t) >= 2]


def _where_like_clause(columns: list[str], tokens: list[str]) -> tuple[str, list[str]]:
    """Costruisce WHERE con AND tra token e OR tra colonne.
    Ogni token deve matchare in almeno una colonna (AND tra token, OR tra colonne)."""
    if not tokens:
        return "1=1", []
    token_groups = []
    params = []
    for token in tokens:
        like = f"%{token}%"
        col_conditions = [f"{col} LIKE ?" for col in columns]
        params.extend([like] * len(columns))
        token_groups.append(f"({' OR '.join(col_conditions)})")
    return " AND ".join(token_groups), params


def search_all(term: str, limit: int = 100) -> dict:
    """Ricerca incrociata su tutti i dataset locali.
    La query viene tokenizzata: ogni token cerca in colonne rilevanti.
    """
    conn = get_conn()
    tokens = _tokenize(term)
    if not tokens:
        conn.close()
        return {"internati": [], "menzioni": [], "decorati": [], "caduti": [], "documenti": [], "fonti_narrative": [], "lettere_personali": [], "term": term, "tokens": []}

    per_table = max(10, limit // 7)
    results = {"internati": [], "menzioni": [], "decorati": [], "caduti": [], "documenti": [], "fonti_narrative": [], "lettere_personali": [], "term": term, "tokens": tokens}

    # ─── IMI internati ───
    cols_i = ["cognome", "nome", "luogo_nascita", "residenza", "luogo_internamento", "grado", "matricola", "sorte", "raw_text"]
    where, params = _where_like_clause(cols_i, tokens)
    rows = conn.execute(
        f"""SELECT id, cognome, nome, data_nascita, luogo_nascita, residenza, grado, luogo_internamento, sorte, lettera, pagina, matricola
            FROM internati WHERE {where} ORDER BY cognome, nome LIMIT ?""",
        params + [per_table],
    ).fetchall()
    results["internati"] = [{"table": "internati", **dict(r)} for r in rows]

    # ─── Menzioni (fondi SME) ───
    cols_m = ["m.cognome", "m.nome", "m.grado", "m.reparto", "m.luogo", "m.data", "m.contesto", "m.testo_originale"]
    where, params = _where_like_clause(cols_m, tokens)
    rows = conn.execute(
        f"""SELECT m.id, m.cognome, m.nome, m.grado, m.reparto, m.luogo, m.data, m.contesto, m.tipo, m.file_pdf, m.pagina, f.codice_fondo, f.titolo
            FROM menzioni m LEFT JOIN fondi_archivistici f ON m.fondo_id = f.id
            WHERE {where} ORDER BY m.cognome, m.nome LIMIT ?""",
        params + [per_table],
    ).fetchall()
    results["menzioni"] = [{"table": "menzioni", "source": "fondi SME", **dict(r)} for r in rows]

    # ─── Decorati ISTORECO ───
    cols_d = ["cognome", "nome", "comune_nascita", "comune_residenza", "data_nascita", "data_morte", "guerra", "grado", "corpo_militare", "reparto", "decorazione", "motivazione", "causa_morte", "luogo_morte", "note"]
    where, params = _where_like_clause(cols_d, tokens)
    rows = conn.execute(
        f"""SELECT id, cognome, nome, comune_nascita, comune_residenza, data_nascita, data_morte, guerra, grado, corpo_militare, reparto, decorazione, motivazione, causa_morte, luogo_morte, albo_nome, url_scheda
            FROM decorati WHERE {where} ORDER BY cognome, nome LIMIT ?""",
        params + [per_table],
    ).fetchall()
    results["decorati"] = [{"table": "decorati", "source": "ISTORECO", **dict(r)} for r in rows]

    # ─── Caduti (tabelle aggregate) ───
    caduti_tables = [
        ("caduti_albooro", ["nominativo", "paternita", "comune_attuale", "grado", "reparto", "luogo_morte", "causa_morte"], "Albi d'Oro"),
        ("caduti_bologna", ["nome", "paternita", "grado", "reparto", "luogo_nascita", "luogo_dimora", "luogo_morte", "decorazioni"], "Bologna"),
        ("caduti_cwgc", ["nome", "cognome", "rank", "service_number", "service", "regiment", "cimitero", "paese_cimitero", "memorial", "unit_detail"], "CWGC"),
        ("caduti_ministero", ["cognome", "nome", "nominativo_paternita", "paternita", "maternita", "comune_nascita", "luogo_sepoltura"], "Ministero Difesa"),
        ("caduti_sardi", ["cognome", "nome", "paternita", "luogo_nascita", "comune_residenza", "grado", "reparto", "luogo_morte", "decorazioni"], "Sardi"),
        ("caduti_francia_ww1", ["nom", "grade", "unite", "lieu_naissance", "bureau_recrutement", "lieu_deces", "pays_deces"], "Francia WW1"),
    ]
    for t, cols, source in caduti_tables:
        try:
            where, params = _where_like_clause([f"{t}.{c}" for c in cols], tokens)
            rows = conn.execute(f"SELECT * FROM {t} WHERE {where} ORDER BY id LIMIT ?", params + [per_table]).fetchall()
            for r in rows:
                d = dict(r)
                d["_source_table"] = t
                d["_source_label"] = source
                d["table"] = "caduti"
                results["caduti"].append(d)
        except Exception:
            pass

    # ─── Decorati Nastro Azzurro ───
    try:
        where, params = _where_like_clause(["cognome", "nome", "arma", "tipo_decorazione"], tokens)
        rows = conn.execute(
            f"""SELECT id, cognome, nome, anno_decorazione, tipo_decorazione, arma, source_id
                FROM decorati_nastroazzurro WHERE {where} ORDER BY cognome, nome LIMIT ?""",
            params + [per_table],
        ).fetchall()
        for r in rows:
            d = dict(r)
            d["table"] = "decorati"
            d["source"] = "Nastro Azzurro"
            results["decorati"].append(d)
    except Exception:
        pass

    # ─── Documenti NARA (ricerca full-text se indicizzato) ───
    try:
        where, params = _where_like_clause(["title", "description", "testo_ocr", "mittente", "destinatario", "unita_citate", "luoghi_citati"], tokens)
        q = f"""
            SELECT id, 'nara_t315' as doc_type, file_immagine as title, NULL as description, testo_ocr as text, data_documento as date, NULL as url, 'NARA T315' as source
            FROM documenti_nara_t315 WHERE {where}
            UNION ALL
            SELECT id, 'nara_catalog' as doc_type, title, description, NULL as text, NULL as date, source_url as url, 'NARA Catalog' as source
            FROM documenti_nara_catalog WHERE {where}
            ORDER BY id LIMIT ?
        """
        rows = conn.execute(q, params + [per_table]).fetchall()
        results["documenti"] = [{"table": "documenti", "doc_type": r["doc_type"], "title": r["title"], "description": r["description"], "text": r["text"], "date": r["date"], "url": r["url"], "source": r["source"]} for r in rows]
    except Exception:
        pass

    # ─── Fonti narrative personali ───
    try:
        where, params = _where_like_clause(["persone_possibili", "titolo", "descrizione", "testo_ocr", "autore", "archivio"], tokens)
        rows = conn.execute(
            f"""SELECT id, nome_file, formato, tipo_fonte, archivio, autore, soggetti_json, persone_possibili,
                       titolo, descrizione, testo_ocr, data_documento, path_locale
                FROM fonti_narrative WHERE {where} ORDER BY data_documento DESC LIMIT ?""",
            params + [per_table],
        ).fetchall()
        results["fonti_narrative"] = [{"table": "fonti_narrative", "source": "Fonte personale", **dict(r)} for r in rows]
    except Exception:
        pass

    # ─── Lettere personali OCR ───
    try:
        where, params = _where_like_clause(["mittente", "destinatario", "luogo", "oggetto", "corpo_testo", "note", "filename"], tokens)
        rows = conn.execute(
            f"""SELECT id, filename, file_path, mittente, destinatario, data_lettera, luogo, oggetto,
                       SUBSTR(corpo_testo, 1, 500) as excerpt, lingua, confidenza, elaborato_il
                FROM lettere_personali WHERE {where} ORDER BY data_lettera DESC, id DESC LIMIT ?""",
            params + [per_table],
        ).fetchall()
        results["lettere_personali"] = [{"table": "lettere_personali", "source": "Lettera personale", **dict(r)} for r in rows]
    except Exception:
        pass

    conn.close()
    return results


def get_fondi_summary() -> list[dict]:
    conn = get_conn()
    rows = conn.execute(
        """SELECT f.file_pdf, f.codice_fondo, MAX(f.titolo) as titolo,
                  COUNT(DISTINCT f.id) as schede,
                  (SELECT COUNT(*) FROM menzioni m WHERE m.file_pdf = f.file_pdf) as menzioni
           FROM fondi_archivistici f GROUP BY f.file_pdf ORDER BY f.file_pdf"""
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def save_internato(lettera: str, file_pdf: str, pagina: int, data: dict, raw_text: str = ""):
    conn = get_conn()
    cols = ["lettera", "file_pdf", "pagina"] + COLUMNS + ["raw_text", "elaborato_il", "needs_review", "review_reason"]
    needs_review = 1 if data.get("needs_review") else 0
    review_reason = data.get("review_reason", "")
    vals = [lettera, file_pdf, pagina] + [data.get(c) for c in COLUMNS] + [raw_text, datetime.now().isoformat(), needs_review, review_reason]
    placeholders = ", ".join(["?"] * len(cols))
    conn.execute(f"INSERT INTO internati ({', '.join(cols)}) VALUES ({placeholders})", vals)
    conn.commit()
    conn.close()


def get_progress(lettera: str) -> Optional[dict]:
    conn = get_conn()
    row = conn.execute("SELECT * FROM progress WHERE lettera = ?", (lettera,)).fetchone()
    conn.close()
    return dict(row) if row else None


def init_progress(lettera: str, total_pages: int):
    conn = get_conn()
    existing = conn.execute("SELECT 1 FROM progress WHERE lettera = ?", (lettera,)).fetchone()
    if existing:
        conn.execute(
            "UPDATE progress SET total_pages = ?, status = 'processing', started_at = ? WHERE lettera = ?",
            (total_pages, datetime.now().isoformat(), lettera),
        )
    else:
        conn.execute(
            "INSERT INTO progress (lettera, total_pages, processed_pages, status, started_at) VALUES (?, ?, 0, 'processing', ?)",
            (lettera, total_pages, datetime.now().isoformat()),
        )
    conn.commit()
    conn.close()


def update_progress(lettera: str, processed: int, status: str = "processing"):
    conn = get_conn()
    conn.execute(
        "UPDATE progress SET processed_pages = ?, status = ? WHERE lettera = ?",
        (processed, status, lettera),
    )
    conn.commit()
    conn.close()


def finish_progress(lettera: str):
    conn = get_conn()
    conn.execute(
        "UPDATE progress SET status = 'done', finished_at = ? WHERE lettera = ?",
        (datetime.now().isoformat(), lettera),
    )
    conn.commit()
    conn.close()


def get_latest_active_progress(statuses: tuple[str, ...] = ("processing", "stopped")) -> Optional[dict]:
    conn = get_conn()
    placeholders = ",".join("?" for _ in statuses)
    row = conn.execute(
        f"SELECT * FROM progress WHERE status IN ({placeholders}) ORDER BY started_at DESC LIMIT 1",
        statuses,
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_all_progress() -> list[dict]:
    conn = get_conn()
    rows = conn.execute("SELECT * FROM progress ORDER BY lettera").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def count_internati() -> int:
    conn = get_conn()
    row = conn.execute("SELECT COUNT(*) as c FROM internati").fetchone()
    conn.close()
    return row["c"]


def count_needs_review() -> int:
    conn = get_conn()
    row = conn.execute("SELECT COUNT(*) as c FROM internati WHERE needs_review = 1").fetchone()
    conn.close()
    return row["c"]


def get_internati(limit: int = 100, offset: int = 0, lettera: str = None, needs_review_only: bool = False) -> list[dict]:
    conn = get_conn()
    query = "SELECT * FROM internati"
    params = []
    conditions = []
    if lettera:
        conditions.append("lettera = ?")
        params.append(lettera)
    if needs_review_only:
        conditions.append("needs_review = 1")
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY id LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def update_internato(rid: int, data: dict) -> bool:
    conn = get_conn()
    sets = []
    vals = []
    for c in COLUMNS:
        if c in data:
            sets.append(f"{c} = ?")
            vals.append(data[c])
    if "needs_review" in data:
        sets.append("needs_review = ?")
        vals.append(1 if data["needs_review"] else 0)
    if "review_reason" in data:
        sets.append("review_reason = ?")
        vals.append(data["review_reason"])
    if "luogo_validato" in data:
        sets.append("luogo_validato = ?")
        vals.append(1 if data["luogo_validato"] else 0)
    if not sets:
        conn.close()
        return False
    vals.append(rid)
    cur = conn.execute(f"UPDATE internati SET {', '.join(sets)} WHERE id = ?", vals)
    conn.commit()
    updated = cur.rowcount > 0
    conn.close()
    return updated


def delete_internato(rid: int) -> bool:
    conn = get_conn()
    cur = conn.execute("DELETE FROM internati WHERE id = ?", (rid,))
    conn.commit()
    deleted = cur.rowcount > 0
    conn.close()
    return deleted


def get_internato_by_id(rid: int) -> Optional[dict]:
    conn = get_conn()
    row = conn.execute("SELECT * FROM internati WHERE id = ?", (rid,)).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_letter(lettera: str) -> int:
    conn = get_conn()
    cur = conn.execute("DELETE FROM internati WHERE lettera = ?", (lettera,))
    conn.execute("DELETE FROM progress WHERE lettera = ?", (lettera,))
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return deleted


def is_page_processed(lettera: str, pagina: int) -> bool:
    conn = get_conn()
    row = conn.execute(
        "SELECT 1 FROM internati WHERE lettera = ? AND pagina = ? LIMIT 1",
        (lettera, pagina),
    ).fetchone()
    conn.close()
    return row is not None


# ─── Entità e collegamenti ───

def _normalize_name(name: str) -> str:
    if not name:
        return ""
    return re.sub(r"\s+", " ", name.strip().lower())


def save_entita(tipo: str, valore: str, fonte_tabella: str, fonte_id: int,
                cognome: str = None, nome: str = None, data: str = None,
                luogo: str = None, contesto: str = None) -> int:
    conn = get_conn()
    norm = _normalize_name(valore)
    existing = conn.execute(
        "SELECT id FROM entita WHERE tipo = ? AND valore_normalizzato = ? LIMIT 1",
        (tipo, norm),
    ).fetchone()
    if existing:
        eid = existing["id"]
    else:
        cur = conn.execute(
            """INSERT INTO entita (tipo, valore, valore_normalizzato, cognome, nome, data, luogo,
               contesto, fonte_tabella, fonte_id, elaborato_il)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (tipo, valore, norm, cognome, nome, data, luogo, contesto,
             fonte_tabella, fonte_id, datetime.now().isoformat()),
        )
        eid = cur.lastrowid
    existing_link = conn.execute(
        """SELECT 1 FROM collegamenti
           WHERE entita_id = ? AND tabella_origine = ? AND record_id = ? LIMIT 1""",
        (eid, fonte_tabella, fonte_id),
    ).fetchone()
    if not existing_link:
        conn.execute(
            """INSERT INTO collegamenti (entita_id, tabella_origine, record_id, tipo_collegamento, elaborato_il)
               VALUES (?, ?, ?, ?, ?)""",
            (eid, fonte_tabella, fonte_id, tipo, datetime.now().isoformat()),
        )
    conn.commit()
    conn.close()
    return eid


def count_entita() -> int:
    conn = get_conn()
    row = conn.execute("SELECT COUNT(*) as c FROM entita").fetchone()
    conn.close()
    return row["c"]


def count_collegamenti() -> int:
    conn = get_conn()
    row = conn.execute("SELECT COUNT(*) as c FROM collegamenti").fetchone()
    conn.close()
    return row["c"]


def search_entita(term: str, limit: int = 50) -> list[dict]:
    conn = get_conn()
    like = f"%{term}%"
    rows = conn.execute(
        """SELECT e.id, e.tipo, e.valore, e.cognome, e.nome, e.data, e.luogo, e.contesto,
                  COUNT(c.id) as num_collegamenti,
                  GROUP_CONCAT(DISTINCT c.tabella_origine) as tabelle
           FROM entita e LEFT JOIN collegamenti c ON e.id = c.entita_id
           WHERE e.valore LIKE ? OR e.cognome LIKE ? OR e.nome LIKE ? OR e.luogo LIKE ?
           GROUP BY e.id ORDER BY num_collegamenti DESC LIMIT ?""",
        (like, like, like, like, limit),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_collegamenti_entita(entita_id: int) -> dict:
    conn = get_conn()
    entita = conn.execute("SELECT * FROM entita WHERE id = ?", (entita_id,)).fetchone()
    if not entita:
        conn.close()
        return None
    colls = conn.execute(
        "SELECT * FROM collegamenti WHERE entita_id = ?", (entita_id,)
    ).fetchall()
    risultati = []
    for c in colls:
        tabella = c["tabella_origine"]
        rid = c["record_id"]
        try:
            row = conn.execute(
                f"SELECT * FROM {tabella} WHERE id = ?", (rid,)
            ).fetchone()
        except sqlite3.OperationalError:
            row = None
        if row:
            risultati.append({"tabella": tabella, "record": dict(row)})
    conn.close()
    return {"entita": dict(entita), "collegamenti": risultati}


def get_all_records_for_ai(term: str, limit_per_table: int = 20) -> dict:
    """Recupera record da tutte le tabelle per fornire contesto alle AI.
    Usa ricerca tokenizzata per query multi-termine."""
    conn = get_conn()
    tokens = _tokenize(term)
    if not tokens:
        conn.close()
        return {"internati": [], "decorati": [], "menzioni": [], "fondi_archivistici": [], "entita": [], "caduti": [], "documenti": [], "fonti_narrative": [], "lettere_personali": [], "term": term, "tokens": []}

    # IMI internati
    cols_i = ["cognome", "nome", "data_nascita", "luogo_nascita", "residenza", "grado", "luogo_cattura", "luogo_internamento", "sorte", "data", "lettera", "pagina", "matricola"]
    where, params = _where_like_clause(cols_i, tokens)
    internati = conn.execute(
        f"""SELECT id, cognome, nome, data_nascita, luogo_nascita, residenza, grado,
                  luogo_cattura, luogo_internamento, sorte, data, lettera, pagina, matricola
           FROM internati WHERE {where} LIMIT ?""",
        params + [limit_per_table],
    ).fetchall()

    # Decorati ISTORECO
    cols_d = ["cognome", "nome", "comune_nascita", "grado", "corpo_militare", "reparto", "decorazione", "guerra", "luogo_morte", "luogo_internamento", "data_nascita", "data_morte", "motivazione", "note"]
    where, params = _where_like_clause(cols_d, tokens)
    decorati = conn.execute(
        f"""SELECT id, cognome, nome, comune_nascita, grado, corpo_militare, reparto,
                  decorazione, guerra, luogo_morte, luogo_internamento, data_nascita, data_morte
           FROM decorati WHERE {where} LIMIT ?""",
        params + [limit_per_table],
    ).fetchall()

    # Menzioni
    cols_m = ["m.cognome", "m.nome", "m.grado", "m.reparto", "m.luogo", "m.data", "m.contesto", "m.testo_originale"]
    where, params = _where_like_clause(cols_m, tokens)
    menzioni = conn.execute(
        f"""SELECT m.id, m.cognome, m.nome, m.grado, m.reparto, m.luogo, m.data, m.contesto,
                  f.codice_fondo, f.titolo
           FROM menzioni m LEFT JOIN fondi_archivistici f ON m.fondo_id = f.id
           WHERE {where} LIMIT ?""",
        params + [limit_per_table],
    ).fetchall()

    # Fondi
    where, params = _where_like_clause(["titolo", "descrizione", "luoghi", "codice_fondo"], tokens)
    fondi = conn.execute(
        f"""SELECT id, codice_fondo, titolo, descrizione, periodo, busta, fascicolo, luoghi
           FROM fondi_archivistici WHERE {where} LIMIT ?""",
        params + [limit_per_table],
    ).fetchall()

    # Entita
    where, params = _where_like_clause(["e.valore", "e.cognome", "e.nome", "e.luogo", "e.data"], tokens)
    entita = conn.execute(
        f"""SELECT e.id, e.tipo, e.valore, e.cognome, e.nome, e.data, e.luogo, e.contesto,
                  COUNT(c.id) as num_collegamenti
           FROM entita e LEFT JOIN collegamenti c ON e.id = c.entita_id
           WHERE {where} GROUP BY e.id LIMIT ?""",
        params + [limit_per_table],
    ).fetchall()

    # Caduti (tabelle aggregate)
    caduti_tables = [
        ("caduti_albooro", ["nominativo", "paternita", "comune_attuale", "grado", "reparto", "luogo_morte", "causa_morte"], "Albi d'Oro"),
        ("caduti_bologna", ["nome", "paternita", "grado", "reparto", "luogo_nascita", "luogo_dimora", "luogo_morte", "decorazioni"], "Bologna"),
        ("caduti_cwgc", ["nome", "cognome", "rank", "service_number", "service", "regiment", "cimitero", "paese_cimitero", "memorial", "unit_detail"], "CWGC"),
        ("caduti_ministero", ["cognome", "nome", "nominativo_paternita", "paternita", "maternita", "comune_nascita", "luogo_sepoltura"], "Ministero Difesa"),
        ("caduti_sardi", ["cognome", "nome", "paternita", "luogo_nascita", "comune_residenza", "grado", "reparto", "luogo_morte", "decorazioni"], "Sardi"),
        ("caduti_francia_ww1", ["nom", "grade", "unite", "lieu_naissance", "bureau_recrutement", "lieu_deces", "pays_deces"], "Francia WW1"),
    ]
    caduti = []
    for t, cols, source in caduti_tables:
        try:
            where, params = _where_like_clause([f"{t}.{c}" for c in cols], tokens)
            rows = conn.execute(f"SELECT * FROM {t} WHERE {where} LIMIT ?", params + [limit_per_table]).fetchall()
            for r in rows:
                d = dict(r); d["_source_label"] = source; d["table"] = "caduti"; caduti.append(d)
        except Exception:
            pass

    # Decorati Nastro Azzurro
    try:
        where, params = _where_like_clause(["cognome", "nome", "arma", "tipo_decorazione"], tokens)
        rows = conn.execute(
            f"""SELECT id, cognome, nome, anno_decorazione, tipo_decorazione, arma
                FROM decorati_nastroazzurro WHERE {where} LIMIT ?""",
            params + [limit_per_table],
        ).fetchall()
        for r in rows:
            d = dict(r); d["source"] = "Nastro Azzurro"; decorati.append(d)
    except Exception:
        pass

    # Documenti NARA
    documenti = []
    try:
        where, params = _where_like_clause(["testo_ocr", "mittente", "destinatario", "unita_citate", "luoghi_citati"], tokens)
        rows = conn.execute(f"SELECT id, file_immagine as title, testo_ocr as text, data_documento as date, 'NARA T315' as source FROM documenti_nara_t315 WHERE {where} LIMIT ?", params + [limit_per_table]).fetchall()
        documenti.extend([dict(r) for r in rows])
    except Exception:
        pass
    try:
        where, params = _where_like_clause(["title", "description"], tokens)
        rows = conn.execute(f"SELECT id, title, description, source_url as url, 'NARA Catalog' as source FROM documenti_nara_catalog WHERE {where} LIMIT ?", params + [limit_per_table]).fetchall()
        documenti.extend([dict(r) for r in rows])
    except Exception:
        pass

    # Fonti narrative personali
    fonti_narrative = []
    try:
        where, params = _where_like_clause(["persone_possibili", "titolo", "descrizione", "testo_ocr", "autore"], tokens)
        rows = conn.execute(
            f"""SELECT id, nome_file, formato, tipo_fonte, archivio, autore, soggetti_json, persone_possibili,
                       titolo, descrizione, testo_ocr, data_documento, path_locale
                FROM fonti_narrative WHERE {where} LIMIT ?""",
            params + [limit_per_table],
        ).fetchall()
        fonti_narrative = [{"table": "fonti_narrative", "source": "Fonte personale", **dict(r)} for r in rows]
    except Exception:
        pass

    # Lettere personali OCR
    lettere_personali = []
    try:
        where, params = _where_like_clause(["mittente", "destinatario", "luogo", "oggetto", "corpo_testo", "note", "filename"], tokens)
        rows = conn.execute(
            f"""SELECT id, filename, file_path, mittente, destinatario, data_lettera, luogo, oggetto,
                       corpo_testo, lingua, confidenza, elaborato_il
                FROM lettere_personali WHERE {where} LIMIT ?""",
            params + [limit_per_table],
        ).fetchall()
        lettere_personali = [{"table": "lettere_personali", "source": "Lettera personale", **dict(r)} for r in rows]
    except Exception:
        pass

    conn.close()
    return {
        "internati": [dict(r) for r in internati],
        "decorati": [dict(r) for r in decorati],
        "menzioni": [dict(r) for r in menzioni],
        "fondi_archivistici": [dict(r) for r in fondi],
        "entita": [dict(r) for r in entita],
        "caduti": caduti,
        "documenti": documenti,
        "fonti_narrative": fonti_narrative,
        "lettere_personali": lettere_personali,
        "term": term,
        "tokens": tokens,
    }


def save_ai_ricerca(query: str, provider: str, model: str, risposta: str,
                    contesto_dati: str, cost_usd: float = 0.0) -> int:
    conn = get_conn()
    cur = conn.execute(
        """INSERT INTO ai_ricerche (query, provider, model, risposta, contesto_dati, cost_usd, elaborato_il)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (query, provider, model, risposta, contesto_dati, cost_usd, datetime.now().isoformat()),
    )
    conn.commit()
    rid = cur.lastrowid
    conn.close()
    return rid


def get_ai_ricerche(limit: int = 20) -> list[dict]:
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM ai_ricerche ORDER BY id DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def export_excel(output_path: str = None) -> str:
    if output_path is None:
        output_path = str(Path(__file__).parent / "export_internati.xlsx")

    conn = get_conn()
    rows = conn.execute("SELECT * FROM internati ORDER BY lettera, id").fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Internati Militari Italiani"

    headers = ["Lettera", "File PDF", "Pagina"] + [c.replace("_", " ").title() for c in COLUMNS]
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    for row_idx, row in enumerate(rows, 2):
        vals = [row["lettera"], row["file_pdf"], row["pagina"]]
        vals += [row[c] or "" for c in COLUMNS]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    col_widths = [8, 25, 8, 20, 20, 15, 20, 20, 10, 20, 15, 20, 20, 20, 20, 15, 20, 30, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path


def export_csv(output_path: str = None) -> str:
    if output_path is None:
        output_path = str(Path(__file__).parent / "export_internati.csv")

    conn = get_conn()
    rows = conn.execute("SELECT * FROM internati ORDER BY lettera, id").fetchall()
    conn.close()

    import csv
    headers = ["lettera", "file_pdf", "pagina"] + COLUMNS
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row["lettera"], row["file_pdf"], row["pagina"]] + [row[c] or "" for c in COLUMNS])

    return output_path

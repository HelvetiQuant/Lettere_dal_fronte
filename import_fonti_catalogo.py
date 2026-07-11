"""Importa il catalogo di 25 fonti storiche da `fonti_scrapabili_metadata.xlsx`
in `fonti_indice` tramite `source_locator.register_source_metadata()`.

Uso:
    python import_fonti_catalogo.py

Requisiti:
    openpyxl
"""
from datetime import datetime
from pathlib import Path

from source_locator import register_source_metadata

EXCEL_PATH = Path(__file__).parent / "fonti_scrapabili_metadata.xlsx"
SHEET_NAME = "fonti_indice"


def _looks_like_url(value: str) -> bool:
    if not value or not isinstance(value, str):
        return False
    value = value.strip()
    return value.startswith(("http://", "https://"))


def _priority_to_confidence(priority: str) -> float:
    return {"alto": 0.8, "medio": 0.6, "basso": 0.4}.get(str(priority).lower().strip(), 0.5)


def _access_type(row: dict) -> str:
    scraping = str(row.get("scraping_html_consentito", "")).lower()
    note = str(row.get("note_legali_tecniche", "")).lower()
    if "solo in loco" in note or "consultazione fisica" in note:
        return "locale"
    if "login" in note or "autenticazione" in note or "partner" in note:
        return "login"
    if scraping.startswith("no") or "viet" in note or "tos" in note:
        return "richiesta"
    return "online"


def _build_note(row: dict) -> str:
    fields = [
        ("dominio", row.get("dominio", "")),
        ("conflitto", row.get("conflitto", "")),
        ("url_pattern_record", row.get("url_pattern_record", "")),
        ("url_esempio_reale_verificato", row.get("url_esempio_reale_verificato", "")),
        ("ha_api", row.get("ha_api", "")),
        ("bulk_download", row.get("bulk_download", "")),
        ("scraping_html_consentito", row.get("scraping_html_consentito", "")),
        ("metodo_accesso_consigliato", row.get("metodo_accesso_consigliato", "")),
        ("priorita_integrazione", row.get("priorita_integrazione", "")),
        ("note_legali_tecniche", row.get("note_legali_tecniche", "")),
        ("verificato_il", row.get("verificato_il", "")),
    ]
    lines = [f"{k}: {v}" for k, v in fields if v not in (None, "")]
    return "\n".join(lines)


def _row_to_meta(row: dict) -> dict:
    nome = str(row.get("nome_fonte", "")).strip()
    categoria = str(row.get("categoria", "")).strip()
    conflitto = str(row.get("conflitto", "")).strip()
    dominio = str(row.get("dominio", "")).strip()

    verified_url = row.get("url_esempio_reale_verificato", "")
    verified_url = verified_url.strip() if isinstance(verified_url, str) else ""

    # Se l'URL verificato e' reale, lo usiamo come url_catalogo;
    # altrimenti puntiamo alla homepage del portale.
    if _looks_like_url(verified_url):
        url_catalogo = verified_url
    else:
        url_catalogo = str(row.get("url_base", "")).strip() or None

    note = _build_note(row)

    return {
        "archivio": nome,
        "segnatura": dominio or nome,
        "titolo": f"{nome} — {categoria}" + (f" ({conflitto})" if conflitto else ""),
        "tipo_fonte": categoria or "fonte_esterna",
        "fondo": str(row.get("tipo_materiale", "")).strip() or None,
        "url_catalogo": url_catalogo,
        "access_type": _access_type(row),
        "confidence": _priority_to_confidence(row.get("priorita_integrazione", "medio")),
        "note": note,
        "last_checked_at": row.get("verificato_il") or datetime.now().isoformat(timespec="seconds"),
    }


def import_catalogo(excel_path: Path = EXCEL_PATH, dry_run: bool = False) -> dict:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl non installato. Installa con: pip install openpyxl") from exc

    if not excel_path.exists():
        raise FileNotFoundError(f"File non trovato: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[SHEET_NAME]

    headers = [c.value for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: v for i, v in enumerate(raw)}
        if not row.get("nome_fonte"):
            continue
        rows.append(row)

    created = updated = skipped = 0
    results = []

    for row in rows:
        meta = _row_to_meta(row)
        if not meta["archivio"]:
            skipped += 1
            continue

        if dry_run:
            results.append({"action": "dry_run", "meta": meta})
            continue

        try:
            res = register_source_metadata(**meta)
            if res.get("created"):
                created += 1
            else:
                updated += 1
            results.append({"id": res.get("id"), "created": res.get("created"), "archivio": meta["archivio"]})
        except Exception as e:
            results.append({"error": str(e), "archivio": meta["archivio"]})
            skipped += 1

    return {
        "total": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "results": results,
    }


def main():
    stats = import_catalogo()
    print(f"Catalogo importato: {stats['total']} fonti")
    print(f"  create:   {stats['created']}")
    print(f"  aggiornate: {stats['updated']}")
    print(f"  saltate:  {stats['skipped']}")


if __name__ == "__main__":
    main()

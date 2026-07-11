import csv
import io
import os
import base64
import json
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
import pdfplumber
import fitz

from config import COLUMNS
from database import save_internato
from extractor import (
    _get_client, _get_mistral_client, _parse_json_response,
    PARSE_PROMPT, _render_page_image, _mistral_ocr_page,
    PARSE_MODEL_MINI, PARSE_MODEL_FULL,
)

UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)


def _normalize_column_name(name: str) -> str:
    n = name.lower().strip()
    n = re.sub(r"[^\w\s]", " ", n)
    n = re.sub(r"\s+", "_", n)
    return n


COLUMN_SYNONYMS = {
    "cognome": "cognome", "surname": "cognome", "last_name": "cognome", "lastname": "cognome",
    "nome": "nome", "name": "nome", "first_name": "nome", "firstname": "nome",
    "data_nascita": "data_nascita", "data_di_nascita": "data_nascita", "birth_date": "data_nascita",
    "birthdate": "data_nascita", "nato": "data_nascita", "nato_il": "data_nascita",
    "luogo_nascita": "luogo_nascita", "luogo_di_nascita": "luogo_nascita",
    "birth_place": "luogo_nascita", "birthplace": "luogo_nascita", "nato_a": "luogo_nascita",
    "residenza": "residenza", "residence": "residenza", "domicilio": "residenza",
    "grado": "grado", "rank": "grado",
    "luogo_cattura": "luogo_cattura", "luogo_di_cattura": "luogo_cattura",
    "capture_place": "luogo_cattura", "cattura": "luogo_cattura",
    "data_cattura": "data_cattura", "data_di_cattura": "data_cattura", "capture_date": "data_cattura",
    "luogo_internamento": "luogo_internamento", "luogo_di_internamento": "luogo_internamento",
    "camp": "luogo_internamento", "lager": "luogo_internamento",
    "matricola": "matricola", "number": "matricola",
    "arbeitskommando": "arbeitskommando", "ak": "arbeitskommando", "kommando": "arbeitskommando",
    "mansione": "mansione", "mansione_svolta_da_internato": "mansione", "job": "mansione", "work": "mansione",
    "sorte": "sorte", "deceduto_disperso_altro": "sorte", "fate": "sorte", "status": "sorte",
    "data": "data", "data_evento": "data", "event_date": "data",
    "documenti": "documenti", "docs": "documenti", "documents": "documenti",
    "riferimento": "documenti", "reference": "documenti",
}


def _suggest_mapping(headers: list) -> dict:
    mapping = {}
    for h in headers:
        norm = _normalize_column_name(str(h))
        if norm in COLUMN_SYNONYMS:
            mapping[str(h)] = COLUMN_SYNONYMS[norm]
        else:
            mapping[str(h)] = None
    return mapping


def import_csv(file_path: Path, column_mapping: dict = None, lettera: str = "IMPORT") -> dict:
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        if column_mapping is None:
            column_mapping = _suggest_mapping(headers)
        count = 0
        skipped = 0
        for row in reader:
            data = {}
            for src_col, db_col in column_mapping.items():
                if db_col and db_col in COLUMNS and src_col in row:
                    val = row[src_col]
                    if val and val.strip():
                        data[db_col] = val.strip()
            if data.get("cognome"):
                save_internato(lettera, file_path.name, 0, data, "")
                count += 1
            else:
                skipped += 1
    return {"imported": count, "skipped": skipped, "headers": list(headers), "suggested_mapping": _suggest_mapping(headers)}


def import_excel(file_path: Path, column_mapping: dict = None, lettera: str = "IMPORT") -> dict:
    wb = load_workbook(str(file_path), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {"imported": 0, "skipped": 0, "headers": [], "suggested_mapping": {}}
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    if column_mapping is None:
        column_mapping = _suggest_mapping(headers)
    count = 0
    skipped = 0
    for row in rows[1:]:
        data = {}
        for i, src_col in enumerate(headers):
            db_col = column_mapping.get(src_col)
            if db_col and db_col in COLUMNS and i < len(row):
                val = row[i]
                if val is not None and str(val).strip():
                    data[db_col] = str(val).strip()
        if data.get("cognome"):
            save_internato(lettera, file_path.name, 0, data, "")
            count += 1
        else:
            skipped += 1
    return {"imported": count, "skipped": skipped, "headers": headers, "suggested_mapping": _suggest_mapping(headers)}


def import_pdf(file_path: Path, lettera: str = "IMPORT", engine: str = "openai") -> dict:
    pdf_doc = pdfplumber.open(str(file_path))
    total = len(pdf_doc.pages)
    count = 0
    for i, page in enumerate(pdf_doc.pages):
        text = ""
        if engine == "mistral":
            try:
                text = _mistral_ocr_page(file_path, i)
            except Exception:
                text = page.extract_text() or ""
        else:
            text = page.extract_text() or ""
        if len(text.strip()) < 50:
            try:
                rows = _parse_image_page_import(file_path, i)
            except Exception:
                rows = []
        else:
            try:
                rows = _parse_text_import(text, i + 1)
            except Exception:
                rows = []
        for row in rows:
            if row.get("cognome"):
                save_internato(lettera, file_path.name, i + 1, row, text[:500])
                count += 1
    pdf_doc.close()
    return {"imported": count, "total_pages": total, "headers": COLUMNS, "suggested_mapping": {c: c for c in COLUMNS}}


def import_image(file_path: Path, lettera: str = "IMPORT") -> dict:
    try:
        client = _get_mistral_client()
        with open(file_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        result = client.ocr.process(
            model="mistral-ocr-latest",
            document={"type": "image_url", "image_url": f"data:image/png;base64,{b64}"},
        )
        text = result.pages[0].markdown if result.pages else ""
    except Exception:
        try:
            with open(file_path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode()
            client = _get_client()
            response = client.chat.completions.create(
                model=PARSE_MODEL_FULL,
                messages=[
                    {"role": "system", "content": PARSE_PROMPT},
                    {"role": "user", "content": [
                        {"type": "text", "text": "Analizza l'immagine ed estrai i dati degli internati."},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "high"}},
                    ]},
                ],
                max_tokens=4096,
                temperature=0.1,
            )
            raw = response.choices[0].message.content.strip()
            rows = _parse_json_response(raw)
            count = 0
            for row in rows:
                if row.get("cognome"):
                    save_internato(lettera, file_path.name, 1, row, "")
                    count += 1
            return {"imported": count, "headers": COLUMNS, "suggested_mapping": {c: c for c in COLUMNS}}
        except Exception as e:
            return {"imported": 0, "error": str(e), "headers": COLUMNS, "suggested_mapping": {}}
    if len(text.strip()) < 20:
        return {"imported": 0, "headers": COLUMNS, "suggested_mapping": {}}
    try:
        rows = _parse_text_import(text, 1)
    except Exception:
        rows = []
    count = 0
    for row in rows:
        if row.get("cognome"):
            save_internato(lettera, file_path.name, 1, row, text[:500])
            count += 1
    return {"imported": count, "headers": COLUMNS, "suggested_mapping": {c: c for c in COLUMNS}}


def _parse_text_import(text: str, page_num: int) -> list:
    client = _get_client()
    response = client.chat.completions.create(
        model=PARSE_MODEL_MINI,
        messages=[
            {"role": "system", "content": PARSE_PROMPT},
            {"role": "user", "content": f"Pagina {page_num}. Testo OCR:\n\n{text}"},
        ],
        max_tokens=4096,
        temperature=0.1,
    )
    raw = response.choices[0].message.content.strip()
    return _parse_json_response(raw)


def _parse_image_page_import(pdf_path: Path, page_num: int) -> list:
    client = _get_client()
    b64 = _render_page_image(pdf_path, page_num)
    response = client.chat.completions.create(
        model=PARSE_MODEL_FULL,
        messages=[
            {"role": "system", "content": PARSE_PROMPT},
            {"role": "user", "content": [
                {"type": "text", "text": f"Pagina {page_num + 1}. Analizza l'immagine ed estrai i dati degli internati."},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "high"}},
            ]},
        ],
        max_tokens=4096,
        temperature=0.1,
    )
    raw = response.choices[0].message.content.strip()
    return _parse_json_response(raw)


def detect_file_type(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".csv": return "csv"
    elif ext in (".xlsx", ".xls"): return "excel"
    elif ext == ".pdf": return "pdf"
    elif ext in (".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".webp"): return "image"
    return "unknown"
